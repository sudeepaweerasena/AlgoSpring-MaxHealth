import openpyxl
import pandas as pd
from datetime import datetime

# File paths and sheet names
OUTPUT_FILE_PATH = "D:\\AlgoSpring\\python\\MaxHealth\\MaxHealth.xlsx"
OUTPUT_SHEET_NAME = "Premium Calculation Sheet"
INPUT_FILE_PATH = "D:\\AlgoSpring\\python\\MaxHealth\\CensusDataNew.xlsx"
INPUT_SHEET_NAME = "Sheet1"
NATIONALITY_UPDATE_SHEET_NAME = "Nationality_Updated"

def nlg_transfer_medical_data():
    try:
        input_df = pd.read_excel(INPUT_FILE_PATH, sheet_name=INPUT_SHEET_NAME)
        print("Initial Input Data:")
        print(input_df.head())  

    except FileNotFoundError:
        print(f"Error: The file {INPUT_FILE_PATH} does not exist.")
        return
    except ValueError:
        print(f"Error: The sheet {INPUT_SHEET_NAME} does not exist in {INPUT_FILE_PATH}.")
        return
    
    column_mapping = {
        "Beneficiary First Name": "Full Name",
        "DOB": "DOB (dd/MM/yyyy)",
        "Marital status": "Marital Status (Single / Married)",
        "Gender": "Gender (M / F or Male / Female)",
        "Relation": "Relation (Employee / Spouse / Child)",
        "Nationality": "Nationality", 
        "Visa Issued Emirates": "Emirate of Visa Issuance",
        "Category": "Category (A-high, B, C, D, E, F)"   
    }

    # category_mapping = {'A': 1, 'B': 2, 'C': 3}

    input_df = input_df.rename(columns=column_mapping)
    print("After Column Mapping:")
    print(input_df.head())  

    try:
        output_wb = openpyxl.load_workbook(OUTPUT_FILE_PATH)
        output_ws = output_wb[OUTPUT_SHEET_NAME]
    except FileNotFoundError:
        print(f"Error: The file {OUTPUT_FILE_PATH} does not exist.")
        return
    except KeyError:
        print(f"Error: The sheet {OUTPUT_SHEET_NAME} does not exist in {OUTPUT_FILE_PATH}.")
        return

    for index, row in input_df.iterrows():
        output_ws.cell(row=index + 2, column=2).value = row.get("Full Name", "")

        dob = row.get("DOB (dd/MM/yyyy)", "")
        if pd.notnull(dob):
            try:
                dob_converted = pd.to_datetime(dob, errors='coerce', format="%d/%m/%Y")
                if pd.notnull(dob_converted):
                    formatted_dob = dob_converted.strftime("%d/%m/%Y")
                else:
                    formatted_dob = "Invalid DOB"
            except ValueError:
                formatted_dob = "Invalid DOB"
            print(f"Processed DOB for {row.get('Full Name', '')}: {formatted_dob}")  # Debugging line
        else:
            formatted_dob = "Invalid DOB"
        output_ws.cell(row=index + 2, column=3).value = formatted_dob

        output_ws.cell(row=index + 2, column=4).value = row.get("Marital Status (Single / Married)", "") 
        output_ws.cell(row=index + 2, column=5).value = row.get("Gender (M / F or Male / Female)", "")
        relation = row.get("Relation (Employee / Spouse / Child)", "")
        output_ws.cell(row=index + 2, column=6).value = relation 
        output_ws.cell(row=index + 2, column=7).value = row.get("Nationality", "")
        output_ws.cell(row=index + 2, column=8).value = row.get("Emirate of Visa Issuance", "")
        output_ws.cell(row=index + 2, column=9).value = "4000"
        output_ws.cell(row=index + 2, column=10).value = row.get("Category (A-high, B, C, D, E, F)", "")
        # category_letter = row.get("Category (A-high, B, C, D, E, F)", "")
        # category_value = category_mapping.get(category_letter, "Unknown")
        # output_ws.cell(row=index + 2, column=10).value = category_value

    output_wb.save(OUTPUT_FILE_PATH)
    print(f"Data successfully written to {OUTPUT_FILE_PATH}")

# # Call the function to execute
# nlg_transfer_medical_data()
